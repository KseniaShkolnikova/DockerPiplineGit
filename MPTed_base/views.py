from datetime import datetime
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.contrib.auth.models import User, Group
from django.contrib.auth import authenticate, login as django_login, logout as django_logout
from django.urls import reverse
from django.views.decorators.http import require_http_methods
from .decorators import custom_login_required, admin_required, student_required
from django.db.models import Q, Count, Avg,  Max, Min
from .utils.email_sender import send_account_changes_email, send_student_credentials_email


# Импортируем твои модели
from api.models import *

@require_http_methods(["GET"])
def login_page(request):
    """HTML страница авторизации"""
    if request.user.is_authenticated:
        if request.user.is_superuser or request.user.groups.filter(name='admin').exists():
            return redirect('admin_dashboard_page')
        return redirect('dashboard_page')  
    return render(request, 'index.html')


@require_http_methods(["POST"])
def login(request):
    """Обработка формы входа (POST запрос)"""
    username = request.POST.get('username', '').strip()
    password = request.POST.get('password', '')
    
    print(f"DEBUG: Попытка входа для пользователя: {username}")
    
    if not username or not password:
        return render(request, 'index.html', {'error': 'Логин и пароль обязательны'})
    
    user = authenticate(username=username, password=password)
    
    if user is not None and user.is_active:
        print(f"DEBUG: Пользователь {username} аутентифицирован")
        
        # Логиним пользователя
        django_login(request, user)
        
        # Определяем куда редиректить
        if user.is_superuser or user.groups.filter(name='admin').exists():
            return redirect('admin_dashboard_page')
        elif user.groups.filter(name='teacher').exists():
            # ПЕРЕНАПРАВЛЯЕМ УЧИТЕЛЕЙ НА ИХ ПОРТАЛ
            return redirect('teacher_portal:dashboard')
        elif user.groups.filter(name='student').exists():
            return redirect('student_dashboard')
        else:
            return redirect('dashboard_page')
    
    print(f"DEBUG: Аутентификация не удалась для {username}")
    return render(request, 'index.html', {'error': 'Неверный логин или пароль'})


@require_http_methods(["GET"])
@custom_login_required
def dashboard_page(request):
    """HTML страница дашборда после авторизации"""
    # Проверяем роль пользователя
    if request.user.is_superuser or request.user.groups.filter(name='admin').exists():
        return redirect('admin_dashboard_page')
    
    # Если пользователь - студент, перенаправляем на студенческий дашборд
    if request.user.groups.filter(name='student').exists():
        return redirect('student_dashboard')
    
    # Если пользователь - учитель, перенаправляем на учительский дашборд
    if request.user.groups.filter(name='teacher').exists():
        # ПЕРЕНАПРАВЛЯЕМ НА УЧИТЕЛЬСКИЙ ПОРТАЛ
        return redirect('teacher_portal:dashboard')  # Используем namespace микросервиса
    
    # Для остальных ролей (или если нет шаблона dashboard.html)
    try:
        return render(request, 'dashboard.html')
    except:
        return render(request, 'index.html', {'error': 'Нет доступных страниц для вашей роли'})


@require_http_methods(["GET"])
@custom_login_required
@admin_required
def admin_dashboard_page(request):
    """HTML страница админ панели"""
    
    # Считаем все данные прямо в view
    total_users = User.objects.count()
    
    # Последние пользователи
    recent_users_qs = User.objects.order_by('-date_joined')[:5]
    
    # Подготавливаем данные для шаблона
    recent_users = []
    for user in recent_users_qs:
        role = 'user'
        if user.is_superuser or user.groups.filter(name='admin').exists():
            role = 'admin'
        elif user.groups.filter(name='teacher').exists():
            role = 'teacher'
        elif user.groups.filter(name='student').exists():
            role = 'student'
        
        recent_users.append({
            'id': user.id,
            'username': user.username,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
            'date_joined': user.date_joined,
            'role': role
        })
    
    # Получаем остальные данные с обработкой исключений
    try:
        total_subjects = Subject.objects.count()
    except:
        total_subjects = 0
    
    try:
        total_groups = StudentGroup.objects.count()
    except:
        total_groups = 0
    
    try:
        active_homework = Homework.objects.filter(is_active=True).count()
    except:
        active_homework = 0
    
    context = {
        'total_users': total_users,
        'recent_users': recent_users,
        'total_subjects': total_subjects,
        'total_groups': total_groups,
        'active_homework': active_homework,
    }
    
    return render(request, 'admin_dashboard.html', context)


@require_http_methods(["GET"])
def logout_view(request):
    """HTML выход из системы"""
    django_logout(request)
    return redirect('/')


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q

# ===== СТРАНИЦЫ УПРАВЛЕНИЯ КЛАССАМИ =====

@custom_login_required
@admin_required
def groups_list(request):
    """Список всех классов"""
    groups = StudentGroup.objects.all().select_related('curator').prefetch_related('students')
    
    # Считаем количество учеников в каждом классе
    for group in groups:
        group.student_count = group.students.count()
    
    context = {
        'groups': groups,
    }
    return render(request, 'admin/groups_list.html', context)


@custom_login_required
@admin_required
def group_create(request):
    """Создание нового класса"""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        year = request.POST.get('year', '').strip()
        curator_id = request.POST.get('curator', '').strip()
        
        # Валидация
        errors = []
        if not name:
            errors.append('Название класса обязательно')
        if not year or not year.isdigit():
            errors.append('Год обучения должен быть числом')
        else:
            year = int(year)
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                # Получаем классного руководителя
                curator = None
                if curator_id:
                    curator = User.objects.get(id=curator_id)
                    # Проверяем, что это учитель
                    if not curator.groups.filter(name='teacher').exists() and not curator.teacher_profile.exists():
                        messages.error(request, 'Классным руководителем может быть только учитель')
                        curator = None
                
                # Создаем класс
                group = StudentGroup.objects.create(
                    name=name,
                    year=year,
                    curator=curator
                )
                
                messages.success(request, f'Класс "{group.name}" успешно создан')
                return redirect('groups_list')
                
            except Exception as e:
                messages.error(request, f'Ошибка при создании класса: {str(e)}')
    
    # Получаем всех учителей для выпадающего списка
    teachers = User.objects.filter(
        Q(groups__name='teacher') | Q(teacher_profile__isnull=False)
    ).distinct().order_by('last_name', 'first_name')
    
    context = {
        'teachers': teachers,
    }
    return render(request, 'admin/group_form.html', context)


@custom_login_required
@admin_required
def group_edit(request, group_id):
    """Редактирование класса"""
    group = get_object_or_404(StudentGroup, id=group_id)
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        year = request.POST.get('year', '').strip()
        curator_id = request.POST.get('curator', '').strip()
        
        # Валидация
        errors = []
        if not name:
            errors.append('Название класса обязательно')
        if not year or not year.isdigit():
            errors.append('Год обучения должен быть числом')
        else:
            year = int(year)
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                # Получаем классного руководителя
                curator = None
                if curator_id:
                    curator = User.objects.get(id=curator_id)
                    # Проверяем, что это учитель
                    if not curator.groups.filter(name='teacher').exists() and not curator.teacher_profile.exists():
                        messages.error(request, 'Классным руководителем может быть только учитель')
                        curator = group.curator  # Оставляем текущего
                
                # Обновляем класс
                group.name = name
                group.year = year
                if curator:
                    group.curator = curator
                group.save()
                
                messages.success(request, f'Класс "{group.name}" успешно обновлен')
                return redirect('groups_list')
                
            except Exception as e:
                messages.error(request, f'Ошибка при обновлении класса: {str(e)}')
    
    # Получаем всех учителей для выпадающего списка
    teachers = User.objects.filter(
        Q(groups__name='teacher') | Q(teacher_profile__isnull=False)
    ).distinct().order_by('last_name', 'first_name')
    
    # Получаем учеников этого класса
    students = StudentProfile.objects.filter(student_group=group).select_related('user')
    
    context = {
        'group': group,
        'teachers': teachers,
        'students': students,
    }
    return render(request, 'admin/group_form.html', context)


@custom_login_required
@admin_required
def group_delete(request, group_id):
    """Удаление класса"""
    group = get_object_or_404(StudentGroup, id=group_id)
    
    if request.method == 'POST':
        group_name = group.name
        group.delete()
        messages.success(request, f'Класс "{group_name}" успешно удален')
        return redirect('groups_list')
    
    # Если GET запрос - редирект на список классов
    return redirect('groups_list')

@custom_login_required
@admin_required
def group_students(request, group_id):
    """Управление учениками в классе"""
    group = get_object_or_404(StudentGroup, id=group_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        student_id = request.POST.get('student_id')
        
        if action == 'add':
            # Добавление ученика в класс
            user = get_object_or_404(User, id=student_id)
            
            # Проверяем, что это ученик
            if not user.groups.filter(name='student').exists():
                return redirect('group_students', group_id=group_id)
            else:
                # Получаем или создаем профиль ученика
                student_profile, created = StudentProfile.objects.get_or_create(
                    user=user,
                    defaults={'course': group.year}  # Устанавливаем курс как у группы
                )
                
                # Проверяем соответствие курса ученика и года группы
                if student_profile.course != group.year:
                    messages.error(
                        request, 
                        f'Нельзя добавить ученика {user.get_full_name()} (курс {student_profile.course}) '
                        f'в группу {group.name} (год {group.year}). '
                        f'Курс ученика должен соответствовать году обучения группы.'
                    )
                else:
                    student_profile.student_group = group
                    student_profile.save()
                    messages.success(request, f'Ученик {user.get_full_name()} добавлен в группу')
        
        elif action == 'remove':
            # Удаление ученика из класса
            student_profile = get_object_or_404(StudentProfile, user_id=student_id)
            student_profile.student_group = None
            student_profile.save()
            messages.success(request, f'Ученик удален из группы')
        
        return redirect('group_students', group_id=group_id)
    
    # Получаем учеников в классе
    students_in_group = StudentProfile.objects.filter(
        student_group=group
    ).select_related('user').order_by('user__last_name', 'user__first_name')
    
    # Получаем учеников без класса, но с соответствующим курсом
    students_without_group = StudentProfile.objects.filter(
        student_group__isnull=True,
        user__groups__name='student',
        course=group.year  # Только ученики с соответствующим курсом
    ).select_related('user').order_by('user__last_name', 'user__first_name')
    
    # Ищем пользователей без профиля, но с ролью student
    users_without_profile = User.objects.filter(
        groups__name='student'
    ).exclude(
        id__in=StudentProfile.objects.values_list('user_id', flat=True)
    ).order_by('last_name', 'first_name')
    
    context = {
        'group': group,
        'students_in_group': students_in_group,
        'students_without_group': students_without_group,
        'users_without_profile': users_without_profile,
    }
    return render(request, 'admin/group_students.html', context)

@custom_login_required
@admin_required
def subjects_list(request):
    """Список всех предметов"""
    subjects = Subject.objects.all().order_by('name')
    
    # Считаем количество учителей и уроков для каждого предмета
    for subject in subjects:
        subject.teacher_count = TeacherSubject.objects.filter(subject=subject).count()
        subject.lesson_count = ScheduleLesson.objects.filter(subject=subject).count()
    
    context = {
        'subjects': subjects,
    }
    return render(request, 'admin/subjects_list.html', context)


@custom_login_required
@admin_required
def subject_create(request):
    """Создание нового предмета"""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        
        # Валидация
        errors = []
        if not name:
            errors.append('Название предмета обязательно')
        elif len(name) < 2:
            errors.append('Название предмета должно быть не менее 2 символов')
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                # Проверяем, нет ли уже предмета с таким названием
                if Subject.objects.filter(name__iexact=name).exists():
                    messages.error(request, 'Предмет с таким названием уже существует')
                else:
                    # Создаем предмет
                    subject = Subject.objects.create(
                        name=name,
                        description=description
                    )
                    return redirect('subjects_list')
                
            except Exception as e:
                messages.error(request, f'Ошибка при создании предмета: {str(e)}')
    
    return render(request, 'admin/subject_form.html')


@custom_login_required
@admin_required
def subject_edit(request, subject_id):
    """Редактирование предмета"""
    subject = get_object_or_404(Subject, id=subject_id)
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        
        # Валидация
        errors = []
        if not name:
            errors.append('Название предмета обязательно')
        elif len(name) < 2:
            errors.append('Название предмета должно быть не менее 2 символов')
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                # Проверяем, нет ли уже предмета с таким названием (кроме текущего)
                if Subject.objects.filter(name__iexact=name).exclude(id=subject_id).exists():
                    messages.error(request, 'Предмет с таким названием уже существует')
                else:
                    # Обновляем предмет
                    subject.name = name
                    subject.description = description
                    subject.save()
                    return redirect('subjects_list')
                
            except Exception as e:
                messages.error(request, f'Ошибка при обновлении предмета: {str(e)}')
    
    context = {
        'subject': subject,
    }
    return render(request, 'admin/subject_form.html', context)


@custom_login_required
@admin_required
def subject_delete(request, subject_id):
    """Удаление предмета"""
    subject = get_object_or_404(Subject, id=subject_id)
    
    if request.method == 'POST':
        # Проверяем, не используется ли предмет
        teacher_count = TeacherSubject.objects.filter(subject=subject).count()
        lesson_count = ScheduleLesson.objects.filter(subject=subject).count()
        
        if teacher_count > 0 or lesson_count > 0:
            messages.error(request, f'Невозможно удалить предмет "{subject.name}", так как он используется ({teacher_count} учителей, {lesson_count} уроков)')
            return redirect('subjects_list')
        
        subject_name = subject.name
        subject.delete()
        return redirect('subjects_list')
    
    # Если GET запрос - редирект на список
    return redirect('subjects_list')

# ===== СТРАНИЦЫ УПРАВЛЕНИЯ УЧИТЕЛЯМИ =====

from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.contrib.auth.hashers import make_password

@custom_login_required
@admin_required
def teachers_list(request):
    """Список всех учителей с поиском и фильтрами"""
    # Получаем всех пользователей с ролью teacher
    teachers_qs = User.objects.filter(
        Q(groups__name='teacher') | Q(teacher_profile__isnull=False)
    ).distinct().order_by('last_name', 'first_name')
    
    # Поиск
    search_query = request.GET.get('search', '').strip()
    if search_query:
        teachers_qs = teachers_qs.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(teacher_profile__patronymic__icontains=search_query)
        ).distinct()
    
    # Фильтрация по статусу аккаунта
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        teachers_qs = teachers_qs.filter(is_active=True)
    elif status_filter == 'inactive':
        teachers_qs = teachers_qs.filter(is_active=False)
    
    # Фильтрация по наличию предметов
    subject_filter = request.GET.get('subject', '')
    if subject_filter:
        if subject_filter == 'with_subjects':
            teachers_qs = teachers_qs.filter(
                teacher_profile__teacher_subjects__isnull=False
            ).distinct()
        elif subject_filter == 'without_subjects':
            teachers_qs = teachers_qs.filter(
                teacher_profile__teacher_subjects__isnull=True
            ).distinct()
    
    # Пагинация
    page_number = request.GET.get('page', 1)
    paginator = Paginator(teachers_qs, 20)  # 20 учителей на странице
    page_obj = paginator.get_page(page_number)
    
    # Подготавливаем данные для шаблона
    teachers = []
    for user in page_obj:
        try:
            profile = user.teacher_profile
            patronymic = profile.patronymic
            phone = profile.phone
            qualification = profile.qualification
        except TeacherProfile.DoesNotExist:
            patronymic = ''
            phone = ''
            qualification = ''
        
        # Получаем предметы учителя
        subjects = Subject.objects.filter(
            subject_teachers__teacher__user=user
        ).order_by('name')
        
        teachers.append({
            'id': user.id,
            'username': user.username,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'patronymic': patronymic,
            'email': user.email,
            'phone': phone,
            'qualification': qualification,
            'is_active': user.is_active,
            'date_joined': user.date_joined,
            'last_login': user.last_login,
            'subject_count': subjects.count(),
            'subjects': subjects[:3],  # Первые 3 предмета для показа
            'all_subjects': list(subjects),  # Все предметы
            'has_profile': hasattr(user, 'teacher_profile'),
        })
    
    # Получаем все предметы для фильтра
    all_subjects = Subject.objects.all().order_by('name')
    
    context = {
        'teachers': teachers,
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'subject_filter': subject_filter,
        'total_count': teachers_qs.count(),
        'active_count': User.objects.filter(
            Q(groups__name='teacher') | Q(teacher_profile__isnull=False),
            is_active=True
        ).distinct().count(),
        'inactive_count': User.objects.filter(
            Q(groups__name='teacher') | Q(teacher_profile__isnull=False),
            is_active=False
        ).distinct().count(),
        'all_subjects': all_subjects,
    }
    return render(request, 'admin/teachers_list.html', context)


@custom_login_required
@admin_required
def teacher_create(request):
    """Создание нового учителя"""
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        confirm_password = request.POST.get('confirm_password', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        patronymic = request.POST.get('patronymic', '').strip()
        phone = request.POST.get('phone', '').strip()
        qualification = request.POST.get('qualification', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        
        # Валидация
        errors = []
        
        if not username:
            errors.append('Имя пользователя обязательно')
        elif User.objects.filter(username=username).exists():
            errors.append('Пользователь с таким именем уже существует')
        
        if email and User.objects.filter(email=email).exists():
            errors.append('Пользователь с таким email уже существует')
        
        if not password:
            errors.append('Пароль обязателен')
        elif len(password) < 6:
            errors.append('Пароль должен быть не менее 6 символов')
        elif password != confirm_password:
            errors.append('Пароли не совпадают')
        
        if not first_name:
            errors.append('Имя обязательно')
        if not last_name:
            errors.append('Фамилия обязательна')
        if not patronymic:
            errors.append('Отчество обязательно')
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                # Создаем пользователя
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    is_active=is_active
                )
                
                # Добавляем в группу teachers
                teacher_group = Group.objects.get(name='teacher')
                user.groups.add(teacher_group)
                
                # Создаем профиль учителя
                TeacherProfile.objects.create(
                    user=user,
                    patronymic=patronymic,
                    phone=phone,
                    qualification=qualification
                )
                
                # ОТПРАВЛЯЕМ EMAIL С УЧЕТНЫМИ ДАННЫМИ УЧИТЕЛЮ
                if email and is_active:
                    full_name = f"{last_name} {first_name} {patronymic}"
                    login_url = request.build_absolute_uri(reverse('login_page'))
                    
                    # Импортируем функцию отправки
                    from .utils.email_sender import send_teacher_credentials_email
                    
                    # Отправляем email
                    email_sent = send_teacher_credentials_email(
                        teacher_email=email,
                        username=username,
                        password=password,
                        teacher_name=full_name,
                        login_url=login_url
                    )
                    
                    if email_sent:
                        messages.success(request, 
                            f'✅ Учитель <strong>{full_name}</strong> успешно создан. '
                            f'<br>📧 Логин и пароль отправлены на email: <strong>{email}</strong>',
                            extra_tags='safe'
                        )
                    else:
                        messages.warning(request, 
                            f'Учитель {full_name} создан, но не удалось отправить email.',
                            extra_tags='warning'
                        )
                else:
                    messages.success(request, f'Учитель {user.get_full_name()} успешно создан')
                
                return redirect('teachers_list')
                
            except Exception as e:
                messages.error(request, f'Ошибка при создании учителя: {str(e)}')
    
    return render(request, 'admin/teacher_form.html')

@custom_login_required
@admin_required
def teacher_edit(request, teacher_id):
    """Редактирование учителя"""
    user = get_object_or_404(User, id=teacher_id)
    
    # Проверяем, что это учитель
    if not user.groups.filter(name='teacher').exists() and not hasattr(user, 'teacher_profile'):
        messages.error(request, 'Пользователь не является учителем')
        return redirect('teachers_list')
    
    try:
        profile = user.teacher_profile
    except TeacherProfile.DoesNotExist:
        profile = None
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        confirm_password = request.POST.get('confirm_password', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        patronymic = request.POST.get('patronymic', '').strip()
        phone = request.POST.get('phone', '').strip()
        qualification = request.POST.get('qualification', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        
        # Валидация
        errors = []
        
        if not username:
            errors.append('Имя пользователя обязательно')
        elif username != user.username and User.objects.filter(username=username).exists():
            errors.append('Пользователь с таким именем уже существует')
        
        if email and email != user.email and User.objects.filter(email=email).exists():
            errors.append('Пользователь с таким email уже существует')
        
        if password:
            if len(password) < 6:
                errors.append('Пароль должен быть не менее 6 символов')
            elif password != confirm_password:
                errors.append('Пароли не совпадают')
        
        if not first_name:
            errors.append('Имя обязательно')
        if not last_name:
            errors.append('Фамилия обязательна')
        if not patronymic:
            errors.append('Отчество обязательно')
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                # Сохраняем старый email для проверки
                old_email = user.email
                email_changed = email != old_email
                
                # Обновляем пользователя
                user.username = username
                user.email = email if email else user.email
                user.first_name = first_name
                user.last_name = last_name
                user.is_active = is_active
                
                # Если изменился пароль
                password_changed = False
                if password:
                    user.password = make_password(password)
                    password_changed = True
                
                user.save()
                
                # Обновляем или создаем профиль
                if profile:
                    profile.patronymic = patronymic
                    profile.phone = phone
                    profile.qualification = qualification
                    profile.save()
                else:
                    TeacherProfile.objects.create(
                        user=user,
                        patronymic=patronymic,
                        phone=phone,
                        qualification=qualification
                    )
                    # Добавляем в группу если не был
                    if not user.groups.filter(name='teacher').exists():
                        teacher_group = Group.objects.get(name='teacher')
                        user.groups.add(teacher_group)
                
                # ОТПРАВЛЯЕМ УВЕДОМЛЕНИЕ ОБ ИЗМЕНЕНИЯХ
                if email and is_active and (email_changed or password_changed):
                    full_name = f"{last_name} {first_name} {patronymic}"
                    login_url = request.build_absolute_uri(reverse('login_page'))
                    
                    # Готовим сообщение об изменениях
                    changes = []
                    if email_changed:
                        changes.append(f"Email изменен на: {email}")
                    if password_changed:
                        changes.append("Пароль был изменен")
                    
                    # Отправляем email об изменениях
                    from .utils.email_sender import send_account_changes_email
                    send_account_changes_email(
                        student_email=email,  # функция универсальная, можно использовать и для учителей
                        username=username,
                        password=password if password_changed else None,
                        student_name=full_name,
                        login_url=login_url,
                        changes=changes
                    )
                
                messages.success(request, f'Данные учителя {user.get_full_name()} успешно обновлены')
                return redirect('teachers_list')
                
            except Exception as e:
                messages.error(request, f'Ошибка при обновлении учителя: {str(e)}')
    
    context = {
        'teacher_user': user,
        'profile': profile,
        'subjects': Subject.objects.filter(subject_teachers__teacher__user=user) if profile else [],
    }
    return render(request, 'admin/teacher_form.html', context)


@custom_login_required
@admin_required
def teacher_toggle_active(request, teacher_id):
    """Блокировка/разблокировка аккаунта учителя"""
    user = get_object_or_404(User, id=teacher_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'toggle':
            user.is_active = not user.is_active
            user.save()
            
            status = "активирован" if user.is_active else "заблокирован"
            messages.success(request, f'Аккаунт учителя {user.get_full_name()} {status}')
        
        return redirect('teachers_list')
    
    # Если GET запрос - редирект на список
    return redirect('teachers_list')


@custom_login_required
@admin_required
def teacher_delete(request, teacher_id):
    """Удаление учителя"""
    user = get_object_or_404(User, id=teacher_id)
    
    if request.method == 'POST':
        # Проверяем, не связан ли учитель с уроками
        lesson_count = ScheduleLesson.objects.filter(teacher=user).count()
        grade_count = Grade.objects.filter(teacher=user).count()
        
        if lesson_count > 0 or grade_count > 0:
            messages.error(request, 
                f'Невозможно удалить учителя "{user.get_full_name()}", так как он ведет уроки '
                f'({lesson_count} уроков) и выставил оценки ({grade_count} оценок)'
            )
            return redirect('teachers_list')
        
        username = user.get_full_name()
        user.delete()
        messages.success(request, f'Учитель {username} успешно удален')
        return redirect('teachers_list')
    
    return redirect('teachers_list')


@custom_login_required
@admin_required
def teacher_subjects(request, teacher_id):
    """Управление предметами учителя"""
    user = get_object_or_404(User, id=teacher_id)
    
    # Проверяем, что это учитель
    if not user.groups.filter(name='teacher').exists():
        messages.error(request, 'Пользователь не является учителем')
        return redirect('teachers_list')
    
    # Получаем или создаем профиль учителя
    profile, created = TeacherProfile.objects.get_or_create(user=user)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        subject_id = request.POST.get('subject_id')
        
        if action == 'add' and subject_id:
            subject = get_object_or_404(Subject, id=subject_id)
            
            # Проверяем, не преподает ли уже этот предмет
            if not TeacherSubject.objects.filter(teacher=profile, subject=subject).exists():
                TeacherSubject.objects.create(teacher=profile, subject=subject)
                messages.success(request, f'Предмет "{subject.name}" добавлен учителю')
        
        elif action == 'remove' and subject_id:
            subject = get_object_or_404(Subject, id=subject_id)
            TeacherSubject.objects.filter(teacher=profile, subject=subject).delete()
            messages.success(request, f'Предмет "{subject.name}" удален у учителя')
        
        return redirect('teacher_subjects', teacher_id=teacher_id)
    
    # Получаем предметы, которые преподает учитель
    teacher_subjects = TeacherSubject.objects.filter(teacher=profile).select_related('subject')
    
    # Получаем все доступные предметы (кроме тех, что уже преподает)
    available_subjects = Subject.objects.exclude(
        id__in=teacher_subjects.values_list('subject_id', flat=True)
    ).order_by('name')
    
    # Получаем расписание учителя
    schedule_lessons = ScheduleLesson.objects.filter(teacher=user).select_related(
        'daily_schedule', 'subject', 'daily_schedule__student_group'
    ).order_by('daily_schedule__week_day', 'lesson_number')[:10]  # Последние 10 уроков
    
    context = {
        'teacher_user': user,
        'profile': profile,
        'teacher_subjects': teacher_subjects,
        'available_subjects': available_subjects,
        'schedule_lessons': schedule_lessons,
        'lesson_count': ScheduleLesson.objects.filter(teacher=user).count(),
    }
    return render(request, 'admin/teacher_subjects.html', context)


@custom_login_required
@admin_required
def teacher_detail(request, teacher_id):
    """Детальная информация об учителе"""
    user = get_object_or_404(User, id=teacher_id)
    
    # Проверяем, что это учитель
    if not user.groups.filter(name='teacher').exists():
        messages.error(request, 'Пользователь не является учителем')
        return redirect('teachers_list')
    
    try:
        profile = user.teacher_profile
    except TeacherProfile.DoesNotExist:
        profile = None
    
    # Получаем предметы учителя
    subjects = TeacherSubject.objects.filter(teacher__user=user).select_related('subject')
    
    # Получаем расписание учителя
    schedule_lessons = ScheduleLesson.objects.filter(teacher=user).select_related(
        'daily_schedule', 'subject', 'daily_schedule__student_group'
    ).order_by('daily_schedule__week_day', 'lesson_number')
    
    # Группируем расписание по дням недели
    schedule_by_day = {}
    for lesson in schedule_lessons:
        day = lesson.daily_schedule.get_week_day_display()
        if day not in schedule_by_day:
            schedule_by_day[day] = []
        schedule_by_day[day].append(lesson)
    
    # Получаем классы, в которых преподает учитель
    teaching_groups = StudentGroup.objects.filter(
        daily_schedules__lessons__teacher=user
    ).distinct().order_by('year', 'name')
    
    # Получаем статистику по оценкам
    grades_given = Grade.objects.filter(teacher=user).count()
    
    context = {
        'teacher_user': user,
        'profile': profile,
        'subjects': subjects,
        'schedule_by_day': schedule_by_day,
        'teaching_groups': teaching_groups,
        'grades_given': grades_given,
        'lesson_count': schedule_lessons.count(),
    }
    return render(request, 'admin/teacher_detail.html', context)


# Дополнение к views.py (добавить в конец)

# ===== СТРАНИЦЫ УПРАВЛЕНИЯ УЧЕНИКАМИ =====

@custom_login_required
@admin_required
def students_list(request):
    """Список всех учеников с поиском и фильтрацией"""
    # Базовый запрос
    students_qs = StudentProfile.objects.select_related(
        'user', 'student_group'
    ).order_by('user__last_name', 'user__first_name')
    
    # Поиск
    search_query = request.GET.get('search', '').strip()
    if search_query:
        students_qs = students_qs.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(patronymic__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(phone__icontains=search_query)
        )
    
    # Фильтрация по классу
    group_filter = request.GET.get('group', '')
    if group_filter:
        if group_filter == 'no_group':
            students_qs = students_qs.filter(student_group__isnull=True)
        else:
            students_qs = students_qs.filter(student_group_id=group_filter)
    
    # Фильтрация по курсу
    course_filter = request.GET.get('course', '')
    if course_filter and course_filter.isdigit():
        students_qs = students_qs.filter(course=int(course_filter))
    
    # Фильтрация по статусу аккаунта
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        students_qs = students_qs.filter(user__is_active=True)
    elif status_filter == 'inactive':
        students_qs = students_qs.filter(user__is_active=False)
    
    # Пагинация
    page_number = request.GET.get('page', 1)
    paginator = Paginator(students_qs, 25)  # 25 учеников на странице
    page_obj = paginator.get_page(page_number)
    
    # Получаем все классы для фильтра
    all_groups = StudentGroup.objects.all().order_by('year', 'name')
    
    # Статистика
    total_students = students_qs.count()
    active_students = students_qs.filter(user__is_active=True).count()
    inactive_students = total_students - active_students
    
    context = {
        'students': page_obj,
        'page_obj': page_obj,
        'all_groups': all_groups,
        'search_query': search_query,
        'group_filter': group_filter,
        'course_filter': course_filter,
        'status_filter': status_filter,
        'total_count': total_students,
        'active_count': active_students,
        'inactive_count': inactive_students,
    }
    return render(request, 'admin/students_list.html', context)


@custom_login_required
@admin_required
def student_detail(request, student_id):
    """Детальная информация об ученике с оценками и расписанием"""
    student_profile = get_object_or_404(StudentProfile, user_id=student_id)
    student_user = student_profile.user
    
    # Получаем все оценки ученика
    grades = Grade.objects.filter(
        student=student_user
    ).select_related('subject', 'teacher', 'schedule_lesson').order_by('-date')
    
    # Рассчитываем средние оценки по предметам
    subject_grades = {}
    for grade in grades:
        subject_name = grade.subject.name
        if subject_name not in subject_grades:
            subject_grades[subject_name] = {
                'subject': grade.subject,
                'grades': [],
                'average': 0
            }
        subject_grades[subject_name]['grades'].append(grade)
    
    # Вычисляем средние
    for subject_data in subject_grades.values():
        grades_list = [float(g.value) for g in subject_data['grades']]
        subject_data['average'] = round(sum(grades_list) / len(grades_list), 1) if grades_list else 0
    
    # Получаем расписание группы ученика (если есть группа)
    schedule_data = []
    if student_profile.student_group:
        daily_schedules = DailySchedule.objects.filter(
            student_group=student_profile.student_group,
            is_active=True,
            is_weekend=False
        ).prefetch_related('lessons__subject').order_by('week_day')
        
        for day_schedule in daily_schedules:
            lessons = day_schedule.lessons.all().order_by('lesson_number')
            schedule_data.append({
                'day': day_schedule.get_week_day_display(),
                'lessons': lessons
            })
    
    # Получаем домашние задания ученика
    homeworks = Homework.objects.filter(
        student_group=student_profile.student_group
    ).order_by('-created_at')[:5] if student_profile.student_group else []
    
    # Статистика
    total_grades = grades.count()
    homework_grades = grades.filter(grade_type='HW').count()
    test_grades = grades.filter(grade_type='TEST').count()
    
    # Последние оценки (10 последних)
    recent_grades = grades[:10]
    
    context = {
        'student_profile': student_profile,
        'student_user': student_user,
        'subject_grades': subject_grades,
        'schedule_data': schedule_data,
        'homeworks': homeworks,
        'total_grades': total_grades,
        'homework_grades': homework_grades,
        'test_grades': test_grades,
        'recent_grades': recent_grades,
        'grades': grades[:20],  # Ограничиваем для производительности
    }
    return render(request, 'admin/student_detail.html', context)


# views.py (в функции student_create)

@custom_login_required
@admin_required
def student_create(request):
    """Создание нового ученика с отправкой email"""
    # Получаем все группы для выпадающего списка
    groups = StudentGroup.objects.all().order_by('year', 'name')
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        confirm_password = request.POST.get('confirm_password', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        patronymic = request.POST.get('patronymic', '').strip()
        phone = request.POST.get('phone', '').strip()
        birth_date = request.POST.get('birth_date', '').strip()
        course = request.POST.get('course', '1').strip()
        address = request.POST.get('address', '').strip()
        group_id = request.POST.get('group', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        
        # Валидация
        errors = []
        
        if not username:
            errors.append('Имя пользователя обязательно')
        elif User.objects.filter(username=username).exists():
            errors.append('Пользователь с таким именем уже существует')
        
        if not email:
            errors.append('Email обязателен для отправки учетных данных')
        elif User.objects.filter(email=email).exists():
            errors.append('Пользователь с таким email уже существует')
        
        if not password:
            errors.append('Пароль обязателен')
        elif len(password) < 6:
            errors.append('Пароль должен быть не менее 6 символов')
        elif password != confirm_password:
            errors.append('Пароли не совпадают')
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                # Создаем пользователя
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    is_active=is_active
                )
                
                # Добавляем в группу students
                student_group_role = Group.objects.get(name='student')
                user.groups.add(student_group_role)
                
                # Получаем учебный класс
                student_group_obj = None
                if group_id and group_id.isdigit():
                    student_group_obj = StudentGroup.objects.filter(id=int(group_id)).first()
                
                # Преобразуем дату рождения
                birth_date_obj = None
                if birth_date:
                    try:
                        birth_date_obj = datetime.strptime(birth_date, '%Y-%m-%d').date()
                    except ValueError:
                        pass
                
                # Создаем профиль ученика
                StudentProfile.objects.create(
                    user=user,
                    patronymic=patronymic,
                    phone=phone,
                    birth_date=birth_date_obj,
                    address=address,
                    course=int(course),
                    student_group=student_group_obj
                )
                
                # Отправляем email с учетными данными
                if email and is_active:
                    full_name = f"{last_name} {first_name} {patronymic}"
                    login_url = request.build_absolute_uri(reverse('login_page'))
                    
                    # Отправляем email
                    email_sent = send_student_credentials_email(
                        student_email=email,
                        username=username,
                        password=password,
                        student_name=full_name,
                        login_url=login_url
                    )
                    
                    if email_sent:
                        messages.success(request, 
                            f'✅ Ученик <strong>{full_name}</strong> успешно создан. '
                            f'<br>📧 Логин и пароль отправлены на email: <strong>{email}</strong>',
                            extra_tags='safe'
                        )
                    else:
                        messages.warning(request, 
                            f'Ученик {full_name} создан, но не удалось отправить email.',
                            extra_tags='warning'
                        )
                else:
                    messages.success(request, f'Ученик {user.get_full_name()} успешно создан')
                
                return redirect('students_list')
                
            except Exception as e:
                messages.error(request, f'❌ Ошибка при создании ученика: {str(e)}')
    
    context = {
        'groups': groups,
    }
    return render(request, 'admin/student_form.html', context)



# views.py (в функции student_edit)
@custom_login_required
@admin_required
def student_edit(request, student_id):
    """Редактирование ученика"""
    student_profile = get_object_or_404(StudentProfile, user_id=student_id)
    student_user = student_profile.user
    
    # Получаем все группы для выпадающего списка
    groups = StudentGroup.objects.all().order_by('year', 'name')
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        confirm_password = request.POST.get('confirm_password', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        patronymic = request.POST.get('patronymic', '').strip()
        phone = request.POST.get('phone', '').strip()
        birth_date = request.POST.get('birth_date', '').strip()
        course = request.POST.get('course', '1').strip()
        address = request.POST.get('address', '').strip()
        group_id = request.POST.get('group', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        
        # Валидация
        errors = []
        
        # ... (существующая валидация)
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                # Сохраняем старый email для проверки
                old_email = student_user.email
                email_changed = email != old_email
                
                # Обновляем пользователя
                student_user.username = username
                student_user.email = email if email else student_user.email
                student_user.first_name = first_name
                student_user.last_name = last_name
                student_user.is_active = is_active
                
                # Если изменился пароль
                password_changed = False
                if password:
                    student_user.password = make_password(password)
                    password_changed = True
                
                student_user.save()
                
                # Если изменился email или пароль, отправляем уведомление
                if email and is_active and (email_changed or password_changed):
                    full_name = f"{last_name} {first_name} {patronymic}"
                    login_url = request.build_absolute_uri(reverse('login_page'))
                    
                    # Готовим сообщение об изменениях
                    changes = []
                    if email_changed:
                        changes.append(f"Email изменен на: {email}")
                    if password_changed:
                        changes.append("Пароль был изменен")
                    
                    # Отправляем email об изменениях
                    send_account_changes_email(
                        student_email=email,
                        username=username,
                        password=password if password_changed else None,
                        student_name=full_name,
                        login_url=login_url,
                        changes=changes
                    )
                
                # Обновляем профиль...
                # ... (существующий код)
                
                messages.success(request, f'Данные ученика {student_user.get_full_name()} успешно обновлены')
                return redirect('students_list')
                
            except Exception as e:
                messages.error(request, f'Ошибка при обновлении ученика: {str(e)}')
    
    context = {
        'student_profile': student_profile,
        'student_user': student_user,
        'groups': groups,
    }
    return render(request, 'admin/student_form.html', context)




@custom_login_required
@admin_required
def student_toggle_active(request, student_id):
    """Блокировка/разблокировка аккаунта ученика"""
    student_profile = get_object_or_404(StudentProfile, user_id=student_id)
    student_user = student_profile.user
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'toggle':
            student_user.is_active = not student_user.is_active
            student_user.save()
            
            status = "активирован" if student_user.is_active else "заблокирован"
            messages.success(request, f'Аккаунт ученика {student_user.get_full_name()} {status}')
        
        return redirect('students_list')
    
    # Если GET запрос - редирект на список
    return redirect('students_list')


@custom_login_required
@admin_required
def student_delete(request, student_id):
    """Удаление ученика"""
    student_profile = get_object_or_404(StudentProfile, user_id=student_id)
    student_user = student_profile.user
    
    if request.method == 'POST':
        # Проверяем, не связан ли ученик с оценками
        grade_count = Grade.objects.filter(student=student_user).count()
        attendance_count = Attendance.objects.filter(student=student_user).count()
        homework_count = HomeworkSubmission.objects.filter(student=student_user).count()
        
        if grade_count > 0 or attendance_count > 0 or homework_count > 0:
            messages.error(request, 
                f'Невозможно удалить ученика "{student_user.get_full_name()}", так как у него есть '
                f'оценки ({grade_count}), посещаемость ({attendance_count}) и домашние задания ({homework_count})'
            )
            return redirect('students_list')
        
        full_name = student_user.get_full_name()
        
        # Удаляем профиль и пользователя
        student_profile.delete()
        student_user.delete()
        
        messages.success(request, f'Ученик {full_name} успешно удален')
        return redirect('students_list')
    
    return redirect('students_list')

# Добавить в views.py

from datetime import datetime, date, timedelta
from django.utils import timezone
from django.db.models import Avg, Count, Sum

# ===== СТРАНИЦЫ ДЛЯ УЧЕНИКОВ =====

# В существующую функцию student_dashboard добавляем получение объявлений

@custom_login_required
@student_required
def student_dashboard(request):
    """Главная страница ученика"""
    today = timezone.now().date()
    
    try:
        student_profile = StudentProfile.objects.get(user=request.user)
    except StudentProfile.DoesNotExist:
        student_profile = StudentProfile.objects.create(
            user=request.user,
            patronymic='',
            course=1
        )
    
    # Только оценки за СЕГОДНЯ
    recent_grades = Grade.objects.filter(
        student=request.user,
        date=today
    ).select_related('subject', 'teacher').order_by('-date')[:10]
    
    # Расписание на сегодня
    today_schedule = []
    current_week_day = ''
    if student_profile.student_group:
        current_week_day = today.strftime('%a').upper()[:3]
        try:
            daily_schedule = DailySchedule.objects.get(
                student_group=student_profile.student_group,
                week_day=current_week_day,
                is_active=True,
                is_weekend=False
            )
            today_schedule = daily_schedule.lessons.select_related(
                'subject', 'teacher'
            ).order_by('lesson_number')
        except DailySchedule.DoesNotExist:
            today_schedule = []
    
    # Домашние задания
    homeworks = Homework.objects.filter(
        student_group=student_profile.student_group,
        due_date__gte=today
    ).select_related('schedule_lesson__subject').order_by('due_date')[:5] if student_profile.student_group else []
    
    # Считаем средний балл и общее количество оценок
    all_grades = Grade.objects.filter(student=request.user)
    total_grades = all_grades.count()
    
    # Вычисляем средний балл
    avg_result = all_grades.aggregate(avg=Avg('value'))
    average_grade = round(avg_result['avg'], 1) if avg_result['avg'] else 0
    
    # Количество предметов
    subject_count = Subject.objects.filter(
        grades__student=request.user
    ).distinct().count()
    
    # === ИСПРАВЛЯЕМ ОШИБКУ: Получаем объявления для ученика ===
    announcements = []
    announcements_count = 0  # Инициализируем счетчик
    
    if student_profile.student_group:
        # Объявления для класса ученика и общие объявления
        announcements = list(Announcement.objects.filter(
            Q(student_group=student_profile.student_group) | Q(is_for_all=True),
            created_at__gte=today - timedelta(days=7)  # За последние 7 дней
        ).select_related('author', 'student_group').order_by('-created_at')[:10])
        
        # Считаем количество объявлений
        announcements_count = len(announcements)  # Используем len() вместо count()
    
    context = {
        'student_profile': student_profile,
        'today_schedule': today_schedule,
        'homeworks': homeworks,
        'recent_grades': recent_grades,
        'today': today,
        'current_week_day': current_week_day,
        'total_grades': total_grades,
        'average_grade': average_grade,
        'subject_count': subject_count,
        # Добавляем объявления в контекст
        'announcements': announcements,
        'announcements_count': announcements_count,  # Передаем число
    }
    return render(request, 'student/dashboard.html', context)

@custom_login_required
@student_required 
def student_schedule(request):
    """Расписание ученика"""
    if not request.user.groups.filter(name='student').exists():
        messages.error(request, 'Доступ только для учеников')
        return redirect('dashboard_page')
    
    try:
        student_profile = StudentProfile.objects.get(user=request.user)
    except StudentProfile.DoesNotExist:
        student_profile = None
    
    # Получаем расписание на всю неделю
    weekly_schedule = []
    if student_profile and student_profile.student_group:
        week_days = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT']
        
        for day_code in week_days:
            try:
                daily_schedule = DailySchedule.objects.get(
                    student_group=student_profile.student_group,
                    week_day=day_code,
                    is_active=True,
                    is_weekend=False
                )
                lessons = daily_schedule.lessons.select_related(
                    'subject', 'teacher'
                ).order_by('lesson_number')
                
                weekly_schedule.append({
                    'day_code': day_code,
                    'day_name': daily_schedule.get_week_day_display(),
                    'lessons': lessons,
                    'is_weekend': daily_schedule.is_weekend,
                    'is_active': daily_schedule.is_active,
                })
            except DailySchedule.DoesNotExist:
                weekly_schedule.append({
                    'day_code': day_code,
                    'day_name': dict(DailySchedule.WeekDay.choices).get(day_code, day_code),
                    'lessons': [],
                    'is_weekend': False,
                    'is_active': False,
                })
    
    context = {
        'student_profile': student_profile,
        'weekly_schedule': weekly_schedule,
    }
    return render(request, 'student/schedule.html', context)


@custom_login_required
@student_required
def student_grades(request):
    """Оценки ученика - таблица по предметам"""
    # Получаем все предметы, по которым есть оценки
    subjects_with_grades = Subject.objects.filter(
        grades__student=request.user
    ).distinct().order_by('name')
    
    # Подготавливаем данные для таблицы
    subject_data = []
    for subject in subjects_with_grades:
        # Получаем все оценки по предмету
        grades = Grade.objects.filter(
            student=request.user,
            subject=subject
        ).order_by('-date')
        
        # Вычисляем средний балл
        average = grades.aggregate(avg=Avg('value'))['avg'] or 0
        
        subject_data.append({
            'subject': subject,
            'grades': grades,
            'average': round(average, 1),
            'count': grades.count(),
        })
    
    # Статистика для круговой диаграммы
    all_grades = Grade.objects.filter(student=request.user)
    total_grades = all_grades.count()
    
    # Группируем оценки по значениям
    grade_stats = {}
    for value in [2, 3, 4, 5]:
        count = all_grades.filter(value=value).count()
        if count > 0:
            percentage = round((count / total_grades) * 100, 1) if total_grades > 0 else 0
            grade_stats[value] = {
                'count': count,
                'percentage': percentage
            }
    
    context = {
        'subject_data': subject_data,
        'total_grades': total_grades,
        'grade_stats': grade_stats,
        'student_profile': request.user.student_profile if hasattr(request.user, 'student_profile') else None,
    }
    return render(request, 'student/grades.html', context)

@custom_login_required
@student_required
def student_homework(request):
    """Домашние задания ученика"""
    try:
        student_profile = StudentProfile.objects.get(user=request.user)
    except StudentProfile.DoesNotExist:
        student_profile = None
    
    # Предметы для фильтра
    subjects = Subject.objects.filter(
        schedule_lessons__daily_schedule__student_group=student_profile.student_group
    ).distinct().order_by('name') if student_profile and student_profile.student_group else Subject.objects.none()
    
    # Фильтры
    status_filter = request.GET.get('status', '')
    subject_filter = request.GET.get('subject', '')
    
    # Базовый запрос
    homeworks_qs = Homework.objects.filter(
        student_group=student_profile.student_group
    ).select_related('schedule_lesson__subject') if student_profile and student_profile.student_group else Homework.objects.none()
    
    homeworks_qs = homeworks_qs.order_by('due_date')
    
    # Применяем фильтры
    if status_filter == 'active':
        homeworks_qs = homeworks_qs.filter(due_date__gte=timezone.now())
    elif status_filter == 'overdue':
        homeworks_qs = homeworks_qs.filter(due_date__lt=timezone.now())
    
    if subject_filter:
        homeworks_qs = homeworks_qs.filter(schedule_lesson__subject_id=subject_filter)
    
    # Получаем отправленные работы
    submissions = HomeworkSubmission.objects.filter(
        student=request.user
    ).select_related('homework')
    
    # Создаем словарь для быстрой проверки
    submission_dict = {sub.homework_id: sub for sub in submissions}
    
    context = {
        'homeworks': homeworks_qs,
        'subjects': subjects,
        'submission_dict': submission_dict,
        'status_filter': status_filter,
        'subject_filter': subject_filter,
        'student_profile': student_profile,
    }
    return render(request, 'student/homework.html', context)

@custom_login_required
@student_required 
def student_attendance(request):
    """Посещаемость ученика"""
    if not request.user.groups.filter(name='student').exists():
        messages.error(request, 'Доступ только для учеников')
        return redirect('dashboard_page')
    
    # Фильтры
    month_filter = request.GET.get('month', '')
    subject_filter = request.GET.get('subject', '')
    
    # Определяем месяц для фильтрации
    today = timezone.now().date()
    if month_filter:
        try:
            year, month = map(int, month_filter.split('-'))
            start_date = date(year, month, 1)
            end_date = date(year, month + 1, 1) if month < 12 else date(year + 1, 1, 1)
        except (ValueError, IndexError):
            start_date = date(today.year, today.month, 1)
            end_date = date(today.year, today.month + 1, 1) if today.month < 12 else date(today.year + 1, 1, 1)
    else:
        start_date = date(today.year, today.month, 1)
        end_date = date(today.year, today.month + 1, 1) if today.month < 12 else date(today.year + 1, 1, 1)
    
    # Получаем посещаемость
    attendance_qs = Attendance.objects.filter(
        student=request.user,
        date__gte=start_date,
        date__lt=end_date
    ).select_related(
        'schedule_lesson__subject',
        'schedule_lesson__daily_schedule'
    ).order_by('-date')
    
    # Фильтрация по предмету
    if subject_filter:
        attendance_qs = attendance_qs.filter(schedule_lesson__subject_id=subject_filter)
    
    # Группируем по дате
    attendance_by_date = {}
    for record in attendance_qs:
        date_str = record.date.strftime('%Y-%m-%d')
        if date_str not in attendance_by_date:
            attendance_by_date[date_str] = {
                'date': record.date,
                'records': []
            }
        attendance_by_date[date_str]['records'].append(record)
    
    # Считаем статистику
    total_lessons = attendance_qs.count()
    present_count = attendance_qs.filter(status='P').count()
    absent_count = attendance_qs.filter(status='A').count()
    late_count = attendance_qs.filter(status='L').count()
    
    # Предметы для фильтра
    subjects = Subject.objects.filter(
        schedule_lessons__attendances__student=request.user
    ).distinct().order_by('name')
    
    # Генерируем список месяцев для выбора
    months = []
    for i in range(6):  # Последние 6 месяцев
        month_date = today - timedelta(days=30*i)
        months.append(month_date.strftime('%Y-%m'))
    
    context = {
        'attendance_by_date': attendance_by_date.values(),
        'total_lessons': total_lessons,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'attendance_rate': round((present_count / total_lessons * 100), 1) if total_lessons > 0 else 0,
        'subjects': subjects,
        'months': months,
        'selected_month': month_filter or today.strftime('%Y-%m'),
        'subject_filter': subject_filter,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'student/attendance.html', context)


@custom_login_required
@student_required 
def student_profile_view(request):
    """Профиль ученика"""
    if not request.user.groups.filter(name='student').exists():
        messages.error(request, 'Доступ только для учеников')
        return redirect('dashboard_page')
    
    try:
        student_profile = StudentProfile.objects.get(user=request.user)
    except StudentProfile.DoesNotExist:
        student_profile = StudentProfile.objects.create(
            user=request.user,
            patronymic='',
            course=1
        )
    
    if request.method == 'POST':
        patronymic = request.POST.get('patronymic', '').strip()
        phone = request.POST.get('phone', '').strip()
        birth_date = request.POST.get('birth_date', '').strip()
        address = request.POST.get('address', '').strip()
        
        # Обновляем профиль
        student_profile.patronymic = patronymic
        student_profile.phone = phone
        
        if birth_date:
            try:
                student_profile.birth_date = datetime.strptime(birth_date, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        student_profile.address = address
        student_profile.save()
        
        messages.success(request, 'Профиль успешно обновлен')
        return redirect('student_profile')
    
    context = {
        'student_profile': student_profile,
    }
    return render(request, 'student/profile.html', context)


@custom_login_required
@student_required 
def student_announcements(request):
    """Объявления для ученика"""
    if not request.user.groups.filter(name='student').exists():
        messages.error(request, 'Доступ только для учеников')
        return redirect('dashboard_page')
    
    try:
        student_profile = StudentProfile.objects.get(user=request.user)
    except StudentProfile.DoesNotExist:
        student_profile = None
    
    # Получаем объявления
    announcements_qs = Announcement.objects.filter(
        Q(student_group=student_profile.student_group) | Q(is_for_all=True)
    ).select_related('author', 'student_group').order_by('-created_at')
    
    # Фильтрация
    group_filter = request.GET.get('group', '')
    if group_filter == 'all':
        announcements_qs = announcements_qs.filter(is_for_all=True)
    elif group_filter and group_filter != 'all':
        announcements_qs = announcements_qs.filter(student_group_id=group_filter)
    
    # Пагинация
    page_number = request.GET.get('page', 1)
    paginator = Paginator(announcements_qs, 15)
    page_obj = paginator.get_page(page_number)
    
    # Группы для фильтра (только те, к которым ученик принадлежит)
    groups = StudentGroup.objects.filter(
        Q(announcements__isnull=False) | Q(students__user=request.user)
    ).distinct().order_by('year', 'name') if student_profile else []
    
    context = {
        'announcements': page_obj,
        'page_obj': page_obj,
        'groups': groups,
        'group_filter': group_filter,
        'student_profile': student_profile,
    }
    return render(request, 'student/announcements.html', context)
# Добавьте в views.py

@require_http_methods(["POST"])
@custom_login_required
@student_required 
def submit_homework(request):
    """Обработка сдачи домашнего задания с файлами"""
    if not request.user.groups.filter(name='student').exists():
        messages.error(request, 'Доступ только для учеников')
        return redirect('student_homework')
    
    homework_id = request.POST.get('homework_id')
    submission_text = request.POST.get('submission_text', '')
    submission_file = request.FILES.get('submission_file')
    
    if not homework_id:
        messages.error(request, 'ID задания не указан')
        return redirect('student_homework')
    
    try:
        homework = Homework.objects.get(id=homework_id)
        
        # Проверяем, что ученик имеет доступ к этому заданию
        student_profile = StudentProfile.objects.get(user=request.user)
        if homework.student_group != student_profile.student_group:
            messages.error(request, 'Доступ к заданию запрещен')
            return redirect('student_homework')
        
        # Проверяем, не сдавал ли уже ученик это задание
        existing_submission = HomeworkSubmission.objects.filter(
            homework=homework, 
            student=request.user
        ).first()
        
        if existing_submission:
            # Если есть существующая отправка, обновляем ее
            if submission_text:
                existing_submission.submission_text = submission_text
            
            if submission_file:
                # Проверяем размер файла
                if submission_file.size > 10 * 1024 * 1024:
                    messages.error(request, 'Файл слишком большой (макс. 10MB)')
                    return redirect('homework_detail', homework_id=homework_id)
                existing_submission.submission_file = submission_file
            
            existing_submission.submitted_at = timezone.now()
            existing_submission.save()
            
            messages.success(request, 'Работа успешно обновлена')
            return redirect('homework_detail', homework_id=homework_id)
        
        # Проверяем срок сдачи
        if homework.due_date < timezone.now():
            messages.error(request, 'Срок сдачи истек')
            return redirect('homework_detail', homework_id=homework_id)
        
        # Проверяем, что хотя бы что-то заполнено
        if not submission_text and not submission_file:
            messages.error(request, 'Заполните текст работы или загрузите файл')
            return redirect('homework_detail', homework_id=homework_id)
        
        # Проверяем размер файла
        if submission_file and submission_file.size > 10 * 1024 * 1024:
            messages.error(request, 'Файл слишком большой (макс. 10MB)')
            return redirect('homework_detail', homework_id=homework_id)
        
        # Создаем новую отправку
        submission = HomeworkSubmission.objects.create(
            homework=homework,
            student=request.user,
            submission_text=submission_text,
            submitted_at=timezone.now()
        )
        
        if submission_file:
            submission.submission_file = submission_file
            submission.save()
        
        messages.success(request, 'Работа успешно сдана')
        return redirect('homework_detail', homework_id=homework_id)
        
    except Homework.DoesNotExist:
        messages.error(request, 'Задание не найдено')
        return redirect('student_homework')
    except StudentProfile.DoesNotExist:
        messages.error(request, 'Профиль ученика не найден')
        return redirect('student_homework')
    except Exception as e:
        messages.error(request, f'Ошибка: {str(e)}')
        return redirect('homework_detail', homework_id=homework_id)
    
@custom_login_required
@student_required
def homework_detail(request, homework_id):
    """Детальная страница домашнего задания"""
    homework = get_object_or_404(Homework, id=homework_id)
    
    # Проверяем, что задание для класса студента
    student_profile = get_object_or_404(StudentProfile, user=request.user)
    if homework.student_group != student_profile.student_group:
        messages.error(request, 'Доступ запрещен')
        return redirect('student_homework')
    
    # Получаем отправку студента (если есть)
    submission = HomeworkSubmission.objects.filter(
        homework=homework,
        student=request.user
    ).first()
    
    context = {
        'homework': homework,
        'submission': submission,
        'student_profile': student_profile,
    }
    return render(request, 'student/homework_detail.html', context)

@require_http_methods(["POST"])
@custom_login_required
@student_required
def delete_submission(request, submission_id):
    """Удаление отправленной работы"""
    submission = get_object_or_404(HomeworkSubmission, id=submission_id, student=request.user)
    
    # Проверяем, можно ли удалить (если срок не истек)
    if submission.homework.due_date < timezone.now():
        messages.error(request, 'Нельзя удалить работу после срока сдачи')
    else:
        submission.delete()
        messages.success(request, 'Работа удалена')
    
    return redirect('student_homework')


from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm

@require_http_methods(["POST"])
@custom_login_required
@student_required
def change_password(request):
    """Смена пароля пользователя"""
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        
        if form.is_valid():
            user = form.save()
            # Обновляем сессию, чтобы пользователь не разлогинился
            update_session_auth_hash(request, user)
            messages.success(request, 'Пароль успешно изменен!')
        else:
            for error in form.errors.values():
                messages.error(request, error[0])
        
        return redirect('student_profile')
    
    return redirect('student_profile')


@custom_login_required
@student_required
def student_profile_view(request):
    """Профиль ученика"""
    # Получаем или создаем профиль ученика
    try:
        student_profile = StudentProfile.objects.get(user=request.user)
    except StudentProfile.DoesNotExist:
        # Создаем профиль с базовыми данными
        student_profile = StudentProfile.objects.create(
            user=request.user,
            patronymic='',
            course=1
        )
        messages.info(request, 'Пожалуйста, заполните свой профиль')
    
    if request.method == 'POST':
        # Получаем данные из формы
        last_name = request.POST.get('last_name', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        patronymic = request.POST.get('patronymic', '').strip()
        phone = request.POST.get('phone', '').strip()
        birth_date = request.POST.get('birth_date', '').strip()
        address = request.POST.get('address', '').strip()
        email = request.POST.get('email', '').strip()
        
        # Валидация обязательных полей
        errors = []
        if not last_name:
            errors.append('Фамилия обязательна')
        if not first_name:
            errors.append('Имя обязательно')
        if not patronymic:
            errors.append('Отчество обязательно')
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                # Обновляем пользователя
                user = request.user
                user.last_name = last_name
                user.first_name = first_name
                
                if email and email != user.email:
                    # Проверяем, не занят ли email
                    if User.objects.filter(email=email).exclude(id=user.id).exists():
                        messages.error(request, 'Этот email уже используется')
                    else:
                        user.email = email
                
                user.save()
                
                # Обновляем профиль ученика
                student_profile.patronymic = patronymic
                student_profile.phone = phone
                student_profile.address = address
                
                if birth_date:
                    try:
                        student_profile.birth_date = datetime.strptime(birth_date, '%Y-%m-%d').date()
                    except ValueError:
                        student_profile.birth_date = None
                else:
                    student_profile.birth_date = None
                
                student_profile.save()
                
                messages.success(request, 'Профиль успешно обновлен')
                return redirect('student_profile')
                
            except Exception as e:
                messages.error(request, f'Ошибка при обновлении профиля: {str(e)}')
    
    context = {
        'student_profile': student_profile,
    }
    return render(request, 'student/profile.html', context)

# ===== АУДИТ ЛОГОВ =====

from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from django.db.models import Q, Count
from datetime import datetime, timedelta

@custom_login_required
@admin_required
def audit_logs(request):
    """Просмотр логов аудита"""
    # Получаем все логи
    logs_qs = AuditLog.objects.select_related('user').order_by('-timestamp')
    
    # Фильтры
    action_filter = request.GET.get('action', '')
    model_filter = request.GET.get('model', '')
    user_filter = request.GET.get('user', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_query = request.GET.get('search', '')
    
    # Применяем фильтры
    if action_filter:
        logs_qs = logs_qs.filter(action=action_filter)
    
    if model_filter:
        logs_qs = logs_qs.filter(model_name=model_filter)
    
    if user_filter and user_filter.isdigit():
        logs_qs = logs_qs.filter(user_id=int(user_filter))
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            logs_qs = logs_qs.filter(timestamp__date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            logs_qs = logs_qs.filter(timestamp__date__lte=date_to_obj)
        except ValueError:
            pass
    
    if search_query:
        logs_qs = logs_qs.filter(
            Q(model_name__icontains=search_query) |
            Q(object_id__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query)
        )
    
    # Статистика
    total_logs = logs_qs.count()
    
    # Статистика по действиям
    action_stats = AuditLog.objects.values('action').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Статистика по моделям
    model_stats = AuditLog.objects.values('model_name').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Пагинация
    from django.core.paginator import Paginator
    page_number = request.GET.get('page', 1)
    paginator = Paginator(logs_qs, 50)  # 50 логов на странице
    page_obj = paginator.get_page(page_number)
    
    # Получаем уникальные значения для фильтров
    users_with_logs = User.objects.filter(
        audit_logs__isnull=False
    ).distinct().order_by('username')
    
    # Подготавливаем данные логов
    logs = []
    for log in page_obj:
        changes_summary = ''
        if log.action == 'CREATE':
            changes_summary = f'Создан объект {log.model_name}'
            if log.new_values:
                # Показываем основные поля
                fields = []
                for key in list(log.new_values.keys())[:3]:  # Показываем первые 3 поля
                    if key not in ['id', 'created_at', 'updated_at']:
                        value = log.new_values.get(key, '')
                        if isinstance(value, str) and len(value) > 50:
                            value = value[:50] + '...'
                        fields.append(f"{key}: {value}")
                if fields:
                    changes_summary = "Поля: " + ", ".join(fields)
        elif log.action == 'UPDATE' and log.old_values and log.new_values:
            changes = []
            for field in log.old_values:
                if field in log.new_values and log.old_values[field] != log.new_values[field]:
                    old_val = str(log.old_values[field])
                    new_val = str(log.new_values[field])
                    
                    # Обрезаем длинные значения
                    if len(old_val) > 30:
                        old_val = old_val[:30] + '...'
                    if len(new_val) > 30:
                        new_val = new_val[:30] + '...'
                    
                    changes.append(f"{field}: {old_val} → {new_val}")
            if changes:
                changes_summary = "Изменения: " + ", ".join(changes[:3])  # Показываем первые 3 изменения
                if len(changes) > 3:
                    changes_summary += f" ...и еще {len(changes) - 3}"
        
        logs.append({
            'id': log.id,
            'user': log.user,
            'action': log.action,
            'action_display': log.get_action_display(),
            'model_name': log.model_name,
            'object_id': log.object_id,
            'changes_summary': changes_summary,
            'timestamp': log.timestamp,
            'ip_address': log.ip_address,
            'user_agent': log.user_agent,
            'request_path': log.request_path,
            'request_method': log.request_method,
            'old_values': log.old_values,
            'new_values': log.new_values,
            'is_system_action': log.is_system_action,
        })
    
    # Определяем доступные модели из логов
    available_models = AuditLog.objects.values_list(
        'model_name', flat=True
    ).distinct().order_by('model_name')
    
    context = {
        'logs': logs,
        'page_obj': page_obj,
        'total_logs': total_logs,
        'action_stats': action_stats,
        'model_stats': model_stats,
        'users_with_logs': users_with_logs,
        'available_models': available_models,
        'action_choices': AuditLog.ActionType.choices,
        'action_filter': action_filter,
        'model_filter': model_filter,
        'user_filter': user_filter,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search_query,
        'today': timezone.now().date(),
        'yesterday': (timezone.now() - timedelta(days=1)).date(),
        'week_ago': (timezone.now() - timedelta(days=7)).date(),
    }
    return render(request, 'admin/audit_logs.html', context)


@custom_login_required
@admin_required
def audit_log_detail(request, log_id):
    """Детальный просмотр записи аудита"""
    log = get_object_or_404(AuditLog, id=log_id)
    
    # Форматируем JSON данные для красивого отображения
    def format_json_data(data):
        if not data:
            return None
        import json
        try:
            return json.dumps(data, indent=2, ensure_ascii=False)
        except:
            return str(data)
    
    context = {
        'log': log,
        'old_values_formatted': format_json_data(log.old_values),
        'new_values_formatted': format_json_data(log.new_values),
        'action_choices': dict(AuditLog.ActionType.choices),
    }
    return render(request, 'admin/audit_log_detail.html', context)


@require_http_methods(["POST"])
@custom_login_required
@admin_required
def clear_audit_logs(request):
    """Очистка старых логов аудита"""
    if request.method == 'POST':
        days_to_keep = request.POST.get('days_to_keep', '90')
        
        try:
            days = int(days_to_keep)
            if days < 1:
                days = 90
            
            # Удаляем логи старше указанного количества дней
            cutoff_date = timezone.now() - timedelta(days=days)
            deleted_count, _ = AuditLog.objects.filter(
                timestamp__lt=cutoff_date
            ).delete()
            
            messages.success(request, f'Удалено {deleted_count} записей аудита старше {days} дней')
        except ValueError:
            messages.error(request, 'Неверное количество дней')
        
        return redirect('admin/audit_logs')
    
    return redirect('admin/audit_logs')


# ===== ПРОСМОТР ОЦЕНОК И СТАТИСТИКИ ПО ГРУППАМ =====



# ===== ПРОСМОТР ИНФОРМАЦИИ ОБ УЧИТЕЛЯХ =====
# В функции student_homework обновите получение информации о файлах:

@custom_login_required
@student_required
def student_homework(request):
    """Домашние задания ученика"""
    try:
        student_profile = StudentProfile.objects.get(user=request.user)
    except StudentProfile.DoesNotExist:
        student_profile = None
    
    # Предметы для фильтра
    subjects = Subject.objects.filter(
        schedule_lessons__daily_schedule__student_group=student_profile.student_group
    ).distinct().order_by('name') if student_profile and student_profile.student_group else Subject.objects.none()
    
    # Фильтры
    status_filter = request.GET.get('status', '')
    subject_filter = request.GET.get('subject', '')
    
    # Базовый запрос
    homeworks_qs = Homework.objects.filter(
        student_group=student_profile.student_group
    ).select_related('schedule_lesson__subject') if student_profile and student_profile.student_group else Homework.objects.none()
    
    homeworks_qs = homeworks_qs.order_by('due_date')
    
    # Применяем фильтры
    if status_filter == 'active':
        homeworks_qs = homeworks_qs.filter(due_date__gte=timezone.now())
    elif status_filter == 'overdue':
        homeworks_qs = homeworks_qs.filter(due_date__lt=timezone.now())
    
    if subject_filter:
        homeworks_qs = homeworks_qs.filter(schedule_lesson__subject_id=subject_filter)
    
    # Получаем отправленные работы
    submissions = HomeworkSubmission.objects.filter(
        student=request.user
    ).select_related('homework')
    
    # Создаем словарь для быстрой проверки
    submission_dict = {sub.homework_id: sub for sub in submissions}
    
    context = {
        'homeworks': homeworks_qs,
        'subjects': subjects,
        'submission_dict': submission_dict,
        'status_filter': status_filter,
        'subject_filter': subject_filter,
        'student_profile': student_profile,
        'today': timezone.now().date(),  # Добавляем сегодняшнюю дату
    }
    return render(request, 'student/homework.html', context)


# Добавьте новую функцию для обработки файлов домашних заданий:

import os
from django.http import FileResponse, HttpResponseForbidden
from django.conf import settings

@custom_login_required
@student_required
def view_homework_file(request, homework_id):
    """Просмотр прикрепленного файла домашнего задания (для учеников)"""
    homework = get_object_or_404(Homework, id=homework_id)
    
    # Проверяем, что ученик имеет доступ к этому заданию
    student_profile = get_object_or_404(StudentProfile, user=request.user)
    if homework.student_group != student_profile.student_group:
        return HttpResponseForbidden("У вас нет доступа к этому файлу")
    
    # Проверяем, есть ли прикрепленный файл
    if not homework.attachment:
        messages.error(request, 'Файл не прикреплен к этому заданию')
        return redirect('student_homework_detail', homework_id=homework_id)
    
    try:
        # Получаем путь к файлу
        file_path = homework.attachment.path
        
        # Проверяем существование файла
        if not os.path.exists(file_path):
            messages.error(request, 'Файл не найден на сервере')
            return redirect('student_homework_detail', homework_id=homework_id)
        
        # Определяем тип файла для правильного Content-Type
        file_extension = os.path.splitext(file_path)[1].lower()
        content_types = {
            '.pdf': 'application/pdf',
            '.doc': 'application/msword',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.txt': 'text/plain',
            '.zip': 'application/zip',
            '.rar': 'application/vnd.rar',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png',
            '.gif': 'image/gif',
        }
        
        content_type = content_types.get(file_extension, 'application/octet-stream')
        
        # Определяем, просмотр или скачивание
        action = request.GET.get('action', 'view')
        
        if action == 'download':
            # Скачивание файла
            response = FileResponse(
                open(file_path, 'rb'),
                content_type=content_type,
                as_attachment=True,
                filename=os.path.basename(file_path)
            )
        else:
            # Просмотр в браузере
            response = FileResponse(
                open(file_path, 'rb'),
                content_type=content_type
            )
            
            # Для изображений добавляем заголовки для правильного отображения
            if file_extension in ['.jpg', '.jpeg', '.png', '.gif']:
                response['Content-Disposition'] = f'inline; filename="{os.path.basename(file_path)}"'
        
        return response
        
    except Exception as e:
        messages.error(request, f'Ошибка при открытии файла: {str(e)}')
        return redirect('student_homework_detail', homework_id=homework_id)


@custom_login_required
@student_required
def view_submission_file(request, submission_id):
    """Просмотр отправленного файла ученика"""
    submission = get_object_or_404(HomeworkSubmission, id=submission_id)
    
    # Проверяем, что это отправка текущего пользователя
    if submission.student != request.user:
        return HttpResponseForbidden("У вас нет доступа к этому файлу")
    
    # Проверяем, есть ли прикрепленный файл
    if not submission.submission_file:
        messages.error(request, 'Файл не прикреплен к этой отправке')
        return redirect('student_homework_detail', homework_id=submission.homework_id)
    
    try:
        # Получаем путь к файлу
        file_path = submission.submission_file.path
        
        # Проверяем существование файла
        if not os.path.exists(file_path):
            messages.error(request, 'Файл не найден на сервере')
            return redirect('student_homework_detail', homework_id=submission.homework_id)
        
        # Определяем тип файла
        file_extension = os.path.splitext(file_path)[1].lower()
        content_types = {
            '.pdf': 'application/pdf',
            '.doc': 'application/msword',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.txt': 'text/plain',
            '.zip': 'application/zip',
            '.rar': 'application/vnd.rar',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png',
            '.gif': 'image/gif',
        }
        
        content_type = content_types.get(file_extension, 'application/octet-stream')
        
        # Определяем, просмотр или скачивание
        action = request.GET.get('action', 'view')
        
        if action == 'download':
            # Скачивание файла
            response = FileResponse(
                open(file_path, 'rb'),
                content_type=content_type,
                as_attachment=True,
                filename=os.path.basename(file_path)
            )
        else:
            # Просмотр в браузере
            response = FileResponse(
                open(file_path, 'rb'),
                content_type=content_type
            )
            
            # Для изображений добавляем заголовки для правильного отображения
            if file_extension in ['.jpg', '.jpeg', '.png', '.gif']:
                response['Content-Disposition'] = f'inline; filename="{os.path.basename(file_path)}"'
        
        return response
        
    except Exception as e:
        messages.error(request, f'Ошибка при открытии файла: {str(e)}')
        return redirect('student_homework_detail', homework_id=submission.homework_id)

# Добавьте в начало файла
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from io import BytesIO
import tempfile
import os
from datetime import datetime

# ===== ИМПОРТ И ЭКСПОРТ УЧЕНИКОВ В EXCEL =====

@custom_login_required
@admin_required
def export_students_excel(request):
    """Экспорт учеников в Excel с учетом текущих фильтров"""
    
    # Получаем те же фильтры, что и в students_list
    search_query = request.GET.get('search', '').strip()
    group_filter = request.GET.get('group', '')
    course_filter = request.GET.get('course', '')
    status_filter = request.GET.get('status', '')
    
    # Базовый запрос с теми же фильтрами
    students_qs = StudentProfile.objects.select_related(
        'user', 'student_group'
    ).order_by('user__last_name', 'user__first_name')
    
    # Применяем фильтры (как в students_list)
    if search_query:
        students_qs = students_qs.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(patronymic__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(phone__icontains=search_query)
        )
    
    if group_filter:
        if group_filter == 'no_group':
            students_qs = students_qs.filter(student_group__isnull=True)
        else:
            students_qs = students_qs.filter(student_group_id=group_filter)
    
    if course_filter and course_filter.isdigit():
        students_qs = students_qs.filter(course=int(course_filter))
    
    if status_filter == 'active':
        students_qs = students_qs.filter(user__is_active=True)
    elif status_filter == 'inactive':
        students_qs = students_qs.filter(user__is_active=False)
    
    # Создаем Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ученики"
    
    # Заголовки
    headers = [
        '№', 'Фамилия', 'Имя', 'Отчество', 'Логин', 'Email',
        'Телефон', 'Курс', 'Класс', 'Дата рождения', 'Адрес',
        'Статус', 'Дата регистрации', 'Последний вход'
    ]
    
    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Применяем стили к заголовкам
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Заполняем данными
    for row_num, student in enumerate(students_qs, 2):
        # Основные данные
        ws.cell(row=row_num, column=1).value = row_num - 1  # №
        ws.cell(row=row_num, column=2).value = student.user.last_name  # Фамилия
        ws.cell(row=row_num, column=3).value = student.user.first_name  # Имя
        ws.cell(row=row_num, column=4).value = student.patronymic or ''  # Отчество
        ws.cell(row=row_num, column=5).value = student.user.username  # Логин
        ws.cell(row=row_num, column=6).value = student.user.email or ''  # Email
        ws.cell(row=row_num, column=7).value = student.phone or ''  # Телефон
        ws.cell(row=row_num, column=8).value = student.course  # Курс
        ws.cell(row=row_num, column=9).value = student.student_group.name if student.student_group else 'Без класса'  # Класс
        ws.cell(row=row_num, column=10).value = student.birth_date.strftime('%d.%m.%Y') if student.birth_date else ''  # Дата рождения
        ws.cell(row=row_num, column=11).value = student.address or ''  # Адрес
        ws.cell(row=row_num, column=12).value = 'Активен' if student.user.is_active else 'Заблокирован'  # Статус
        ws.cell(row=row_num, column=13).value = student.user.date_joined.strftime('%d.%m.%Y %H:%M')  # Дата регистрации
        ws.cell(row=row_num, column=14).value = student.user.last_login.strftime('%d.%m.%Y %H:%M') if student.user.last_login else 'Никогда'  # Последний вход
        
        # Выравнивание
        for col_num in range(1, 15):
            ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="left", vertical="center")
    
    # Автоподбор ширины колонок
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        
        # Получаем максимальную длину в колонке
        max_length = len(header)
        for row_num in range(2, len(students_qs) + 2):
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        # Устанавливаем ширину с запасом
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Добавляем строку с итогами
    total_row = len(students_qs) + 3
    ws.cell(row=total_row, column=1).value = f"Всего учеников: {len(students_qs)}"
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    
    # Создаем HttpResponse с Excel файлом
    filename = f"ucheniki_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Сохраняем в response
    wb.save(response)
    
    return response


@custom_login_required
@admin_required
def export_students_template(request):
    """Скачивание шаблона для импорта учеников"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Шаблон для импорта"
    
    # Заголовки для импорта
    headers = [
        'Фамилия*', 'Имя*', 'Отчество*', 'Логин*', 'Email*', 
        'Пароль*', 'Телефон', 'Курс*', 'Класс', 'Дата рождения (ДД.ММ.ГГГГ)',
        'Адрес', 'Активен (да/нет)'
    ]
    
    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Стиль для обязательных полей (красная звездочка)
    required_font = Font(bold=True, color="FF0000")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Добавляем пояснения
    ws.cell(row=2, column=1).value = "Иванов"
    ws.cell(row=2, column=2).value = "Иван"
    ws.cell(row=2, column=3).value = "Иванович"
    ws.cell(row=2, column=4).value = "ivanov_2024"
    ws.cell(row=2, column=5).value = "ivanov@example.com"
    ws.cell(row=2, column=6).value = "password123"
    ws.cell(row=2, column=7).value = "+7 (999) 123-45-67"
    ws.cell(row=2, column=8).value = "1"
    ws.cell(row=2, column=9).value = "10А"
    ws.cell(row=2, column=10).value = "15.05.2008"
    ws.cell(row=2, column=11).value = "ул. Пушкина, д. 10"
    ws.cell(row=2, column=12).value = "да"
    
    # Выравнивание для примера
    for col_num in range(1, 13):
        ws.cell(row=2, column=col_num).alignment = Alignment(horizontal="left", vertical="center")
    
    # Добавляем лист с инструкцией
    ws_instruction = wb.create_sheet("Инструкция")
    
    instruction_data = [
        ["ИНСТРУКЦИЯ ПО ИМПОРТУ УЧЕНИКОВ"],
        [""],
        ["1. Поля, отмеченные звездочкой (*), обязательны для заполнения"],
        ["2. Логин должен быть уникальным для каждого ученика"],
        ["3. Email должен быть уникальным и корректным"],
        ["4. Пароль должен быть не менее 6 символов"],
        ["5. Курс указывается числом (1, 2, 3 или 4)"],
        ["6. Класс должен существовать в системе (название класса должно точно совпадать)"],
        ["7. Дата рождения в формате: ДД.ММ.ГГГГ (например, 15.05.2008)"],
        ["8. Поле 'Активен' может быть: да/нет, true/false, 1/0"],
        [""],
        ["ПРИМЕР ЗАПОЛНЕНИЯ:"],
        ["Фамилия", "Имя", "Отчество", "Логин", "Email", "Пароль", "Телефон", "Курс", "Класс", "Дата рождения", "Адрес", "Активен"],
        ["Петров", "Петр", "Петрович", "petrov_2024", "petrov@mail.ru", "pass123", "+79991112233", "2", "10Б", "10.08.2007", "ул. Ленина 5", "да"],
    ]
    
    for row_num, row_data in enumerate(instruction_data, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_instruction.cell(row=row_num, column=col_num)
            cell.value = value
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
    
    # Автоподбор ширины для инструкции
    ws_instruction.column_dimensions['A'].width = 60
    
    # Устанавливаем ширину колонок для шаблона
    for col_num in range(1, 13):
        column_letter = get_column_letter(col_num)
        if col_num <= 6:  # Логин, пароль и т.д.
            ws.column_dimensions[column_letter].width = 20
        elif col_num == 10:  # Дата рождения
            ws.column_dimensions[column_letter].width = 18
        elif col_num == 11:  # Адрес
            ws.column_dimensions[column_letter].width = 30
        else:
            ws.column_dimensions[column_letter].width = 15
    
    filename = f"shablon_importa_uchenikov.xlsx"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    
    return response


@require_http_methods(["POST"])
@custom_login_required
@admin_required
def import_students_excel(request):
    """Импорт учеников из Excel файла"""
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        # Проверка расширения файла
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Пожалуйста, загрузите файл формата .xlsx или .xls')
            return redirect('students_list')
        
        try:
            # Загружаем рабочую книгу
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            # Проверяем, что файл не пустой
            if ws.max_row < 2:
                messages.error(request, 'Файл не содержит данных для импорта')
                return redirect('students_list')
            
            # Получаем заголовки для проверки
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    # Убираем звездочку из заголовка для сравнения
                    clean_header = header.replace('*', '').strip()
                    headers.append(clean_header)
            
            # Проверяем наличие обязательных колонок
            required_columns = ['Фамилия', 'Имя', 'Отчество', 'Логин', 'Email', 'Пароль', 'Курс']
            missing_columns = []
            for req in required_columns:
                if req not in headers:
                    missing_columns.append(req)
            
            if missing_columns:
                messages.error(request, f'В файле отсутствуют обязательные колонки: {", ".join(missing_columns)}')
                return redirect('students_list')
            
            # Статистика импорта
            created_count = 0
            updated_count = 0
            error_count = 0
            errors = []
            
            # Получаем группу students
            student_group_role = Group.objects.get(name='student')
            
            # Кэш для классов (чтобы не делать запросы к БД для каждой строки)
            group_cache = {}
            
            # Обрабатываем строки (начиная со 2 строки)
            for row_num in range(2, ws.max_row + 1):
                try:
                    # Пропускаем пустые строки
                    last_name = ws.cell(row=row_num, column=headers.index('Фамилия') + 1).value
                    if not last_name or str(last_name).strip() == '':
                        continue
                    
                    # Собираем данные из строки
                    first_name = ws.cell(row=row_num, column=headers.index('Имя') + 1).value
                    patronymic = ws.cell(row=row_num, column=headers.index('Отчество') + 1).value
                    username = ws.cell(row=row_num, column=headers.index('Логин') + 1).value
                    email = ws.cell(row=row_num, column=headers.index('Email') + 1).value
                    password = ws.cell(row=row_num, column=headers.index('Пароль') + 1).value
                    
                    # Преобразуем значения в строки
                    last_name = str(last_name).strip() if last_name else ''
                    first_name = str(first_name).strip() if first_name else ''
                    patronymic = str(patronymic).strip() if patronymic else ''
                    username = str(username).strip() if username else ''
                    email = str(email).strip() if email else ''
                    password = str(password).strip() if password else ''
                    
                    # Проверка обязательных полей
                    if not all([last_name, first_name, patronymic, username, email, password]):
                        errors.append(f'Строка {row_num}: Заполните все обязательные поля')
                        error_count += 1
                        continue
                    
                    # Проверка длины пароля
                    if len(password) < 6:
                        errors.append(f'Строка {row_num}: Пароль должен быть не менее 6 символов')
                        error_count += 1
                        continue
                    
                    # Курс
                    course_col = headers.index('Курс') + 1
                    course_val = ws.cell(row=row_num, column=course_col).value
                    try:
                        course = int(float(course_val)) if course_val else 1
                        if course not in [1, 2, 3, 4]:
                            course = 1
                    except (ValueError, TypeError):
                        course = 1
                    
                    # Телефон (необязательный)
                    phone = ''
                    if 'Телефон' in headers:
                        phone_col = headers.index('Телефон') + 1
                        phone_val = ws.cell(row=row_num, column=phone_col).value
                        phone = str(phone_val).strip() if phone_val else ''
                    
                    # Класс (необязательный)
                    student_group = None
                    if 'Класс' in headers:
                        group_col = headers.index('Класс') + 1
                        group_name = ws.cell(row=row_num, column=group_col).value
                        if group_name:
                            group_name = str(group_name).strip()
                            # Используем кэш
                            if group_name in group_cache:
                                student_group = group_cache[group_name]
                            else:
                                try:
                                    student_group = StudentGroup.objects.get(name=group_name)
                                    group_cache[group_name] = student_group
                                except StudentGroup.DoesNotExist:
                                    errors.append(f'Строка {row_num}: Класс "{group_name}" не найден в системе')
                                    error_count += 1
                                    continue
                    
                    # Дата рождения
                    birth_date = None
                    if 'Дата рождения (ДД.ММ.ГГГГ)' in headers:
                        birth_col = headers.index('Дата рождения (ДД.ММ.ГГГГ)') + 1
                        birth_val = ws.cell(row=row_num, column=birth_col).value
                        if birth_val:
                            try:
                                if isinstance(birth_val, datetime):
                                    birth_date = birth_val.date()
                                elif isinstance(birth_val, str):
                                    # Пробуем разные форматы
                                    for fmt in ['%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d']:
                                        try:
                                            birth_date = datetime.strptime(birth_val.strip(), fmt).date()
                                            break
                                        except ValueError:
                                            continue
                            except:
                                pass
                    
                    # Адрес
                    address = ''
                    if 'Адрес' in headers:
                        addr_col = headers.index('Адрес') + 1
                        addr_val = ws.cell(row=row_num, column=addr_col).value
                        address = str(addr_val).strip() if addr_val else ''
                    
                    # Активен
                    is_active = True
                    if 'Активен (да/нет)' in headers:
                        active_col = headers.index('Активен (да/нет)') + 1
                        active_val = ws.cell(row=row_num, column=active_col).value
                        if active_val:
                            active_str = str(active_val).lower().strip()
                            if active_str in ['нет', 'no', 'false', '0', '']:
                                is_active = False
                    
                    # Проверяем, существует ли пользователь с таким логином или email
                    user = None
                    is_new = True
                    
                    try:
                        user = User.objects.get(username=username)
                        is_new = False
                        # Обновляем существующего пользователя
                        user.email = email
                        user.first_name = first_name
                        user.last_name = last_name
                        user.is_active = is_active
                        if password and password != '********':
                            user.set_password(password)
                        user.save()
                        
                        # Обновляем профиль
                        profile, created = StudentProfile.objects.get_or_create(
                            user=user,
                            defaults={
                                'patronymic': patronymic,
                                'course': course,
                                'phone': phone,
                                'birth_date': birth_date,
                                'address': address,
                                'student_group': student_group
                            }
                        )
                        if not created:
                            profile.patronymic = patronymic
                            profile.course = course
                            profile.phone = phone
                            profile.birth_date = birth_date
                            profile.address = address
                            profile.student_group = student_group
                            profile.save()
                        
                        updated_count += 1
                        
                    except User.DoesNotExist:
                        # Проверяем email на уникальность
                        if User.objects.filter(email=email).exists():
                            errors.append(f'Строка {row_num}: Email {email} уже используется другим пользователем')
                            error_count += 1
                            continue
                        
                        # Создаем нового пользователя
                        user = User.objects.create_user(
                            username=username,
                            email=email,
                            password=password,
                            first_name=first_name,
                            last_name=last_name,
                            is_active=is_active
                        )
                        
                        # Добавляем в группу students
                        user.groups.add(student_group_role)
                        
                        # Создаем профиль ученика
                        StudentProfile.objects.create(
                            user=user,
                            patronymic=patronymic,
                            phone=phone,
                            birth_date=birth_date,
                            address=address,
                            course=course,
                            student_group=student_group
                        )
                        
                        created_count += 1
                        
                except Exception as e:
                    errors.append(f'Строка {row_num}: Ошибка - {str(e)}')
                    error_count += 1
            
            # Формируем сообщение о результате
            if created_count > 0 or updated_count > 0:
                success_msg = f'✅ Импорт завершен: создано {created_count} учеников, обновлено {updated_count} учеников'
                if error_count > 0:
                    success_msg += f', ошибок: {error_count}'
                messages.success(request, success_msg)
                
                # Если есть ошибки, показываем их
                if errors:
                    error_text = "\n".join(errors[:10])  # Показываем первые 10 ошибок
                    if len(errors) > 10:
                        error_text += f"\n...и еще {len(errors) - 10} ошибок"
                    messages.warning(request, f'Ошибки при импорте:\n{error_text}')
            else:
                messages.error(request, f'Не удалось импортировать данные. Ошибок: {error_count}')
            
        except Exception as e:
            messages.error(request, f'Ошибка при обработке файла: {str(e)}')
        
        return redirect('students_list')
    
    messages.error(request, 'Пожалуйста, выберите файл для импорта')
    return redirect('students_list')
@custom_login_required
@admin_required
def export_groups_excel(request):
    """Экспорт всех групп с учениками в Excel (каждая группа - отдельный лист)"""
    
    # Получаем все группы с учениками
    groups = StudentGroup.objects.all().order_by('year', 'name').prefetch_related(
        'students__user'
    )
    
    # Создаем Excel файл
    wb = openpyxl.Workbook()
    
    # Удаляем стандартный лист (мы создадим свои)
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Если нет групп, создаем информационный лист
    if not groups:
        ws = wb.create_sheet("Информация")
        ws.cell(row=1, column=1).value = "В системе нет групп"
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.column_dimensions['A'].width = 30
    else:
        # Для каждой группы создаем отдельный лист
        for group in groups:
            # Название листа (обрезаем если слишком длинное)
            sheet_name = f"{group.name}"[:31]  # Excel ограничение на длину имени листа
            
            # Проверяем уникальность имени
            base_name = sheet_name
            counter = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{base_name}_{counter}"[:31]
                counter += 1
            
            ws = wb.create_sheet(sheet_name)
            
            # Заголовок группы
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = f"Класс: {group.name} ({group.year} курс)"
            title_cell.font = Font(bold=True, size=14)
            ws.merge_cells('A1:E1')
            title_cell.alignment = Alignment(horizontal="center")
            
            # Информация о классном руководителе
            if group.curator:
                curator_name = group.curator.get_full_name()
                curator_cell = ws.cell(row=2, column=1)
                curator_cell.value = f"Классный руководитель: {curator_name}"
                curator_cell.font = Font(bold=True, italic=True)
                ws.merge_cells('A2:E2')
                curator_cell.alignment = Alignment(horizontal="center")
            
            # Заголовки таблицы
            headers = ['№', 'Фамилия', 'Имя', 'Отчество', 'Статус']
            header_row = 4 if group.curator else 3
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Получаем учеников группы
            students = group.students.all().select_related('user').order_by('user__last_name', 'user__first_name')
            
            # Заполняем данными
            for row_num, student in enumerate(students, header_row + 1):
                ws.cell(row=row_num, column=1).value = row_num - header_row
                ws.cell(row=row_num, column=2).value = student.user.last_name
                ws.cell(row=row_num, column=3).value = student.user.first_name
                ws.cell(row=row_num, column=4).value = student.patronymic or ''
                
                status = 'Активен' if student.user.is_active else 'Заблокирован'
                ws.cell(row=row_num, column=5).value = status
                
                # Выравнивание
                for col_num in range(1, 6):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
                    # Цвет статуса
                    if col_num == 5:
                        if student.user.is_active:
                            cell.font = Font(color="008000")  # Зеленый
                        else:
                            cell.font = Font(color="FF0000")  # Красный
            
            # Если в группе нет учеников
            if not students:
                no_students_cell = ws.cell(row=header_row + 1, column=1)
                no_students_cell.value = "В группе нет учеников"
                no_students_cell.font = Font(italic=True, color="666666")
                ws.merge_cells(start_row=header_row + 1, start_column=1, end_row=header_row + 1, end_column=5)
                no_students_cell.alignment = Alignment(horizontal="center")
            
            # Статистика группы
            stats_row = header_row + len(students) + 3
            ws.cell(row=stats_row, column=1).value = f"Всего учеников: {len(students)}"
            ws.cell(row=stats_row, column=1).font = Font(bold=True)
            ws.merge_cells(start_row=stats_row, start_column=1, end_row=stats_row, end_column=5)
            
            active_count = students.filter(user__is_active=True).count()
            inactive_count = len(students) - active_count
            
            ws.cell(row=stats_row + 1, column=1).value = f"Активных: {active_count}"
            ws.cell(row=stats_row + 1, column=1).font = Font(color="008000")
            ws.merge_cells(start_row=stats_row + 1, start_column=1, end_row=stats_row + 1, end_column=5)
            
            ws.cell(row=stats_row + 2, column=1).value = f"Заблокированных: {inactive_count}"
            ws.cell(row=stats_row + 2, column=1).font = Font(color="FF0000")
            ws.merge_cells(start_row=stats_row + 2, start_column=1, end_row=stats_row + 2, end_column=5)
            
            # Автоподбор ширины колонок
            for col_num in range(1, 6):
                column_letter = get_column_letter(col_num)
                
                max_length = len(headers[col_num-1])
                for row_num in range(header_row + 1, header_row + len(students) + 1):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                
                adjusted_width = min(max_length + 2, 40)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    # Создаем итоговый лист со сводкой
    summary_sheet = wb.create_sheet("Сводка по классам", 0)  # Первый лист
    
    # Заголовок сводки
    summary_sheet.cell(row=1, column=1).value = "СВОДКА ПО КЛАССАМ"
    summary_sheet.cell(row=1, column=1).font = Font(bold=True, size=16)
    summary_sheet.merge_cells('A1:D1')
    summary_sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    
    # Заголовки таблицы сводки
    summary_headers = ['№', 'Класс', 'Курс', 'Классный руководитель', 'Учеников', 'Активных', 'Заблокировано']
    header_row = 3
    
    for col_num, header in enumerate(summary_headers, 1):
        cell = summary_sheet.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Заполняем сводку
    for idx, group in enumerate(groups, 1):
        student_count = group.students.count()
        active_count = group.students.filter(user__is_active=True).count()
        inactive_count = student_count - active_count
        
        summary_sheet.cell(row=header_row + idx, column=1).value = idx
        summary_sheet.cell(row=header_row + idx, column=2).value = group.name
        summary_sheet.cell(row=header_row + idx, column=3).value = f"{group.year} курс"
        summary_sheet.cell(row=header_row + idx, column=4).value = group.curator.get_full_name() if group.curator else "Не назначен"
        summary_sheet.cell(row=header_row + idx, column=5).value = student_count
        summary_sheet.cell(row=header_row + idx, column=6).value = active_count
        summary_sheet.cell(row=header_row + idx, column=7).value = inactive_count
        
        # Выравнивание
        for col_num in range(1, 8):
            cell = summary_sheet.cell(row=header_row + idx, column=col_num)
            cell.alignment = Alignment(horizontal="center" if col_num > 4 else "left", vertical="center")
    
    # Итоги в сводке
    total_row = header_row + len(groups) + 2
    total_students = sum(g.students.count() for g in groups)
    total_active = sum(g.students.filter(user__is_active=True).count() for g in groups)
    total_inactive = total_students - total_active
    
    summary_sheet.cell(row=total_row, column=1).value = "ИТОГО:"
    summary_sheet.cell(row=total_row, column=1).font = Font(bold=True)
    summary_sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    summary_sheet.cell(row=total_row, column=5).value = total_students
    summary_sheet.cell(row=total_row, column=6).value = total_active
    summary_sheet.cell(row=total_row, column=7).value = total_inactive
    
    for col_num in range(5, 8):
        summary_sheet.cell(row=total_row, column=col_num).font = Font(bold=True)
    
    # Автоподбор ширины для сводки
    for col_num, header in enumerate(summary_headers, 1):
        column_letter = get_column_letter(col_num)
        
        max_length = len(header)
        for row_num in range(header_row + 1, header_row + len(groups) + 1):
            cell_value = summary_sheet.cell(row=row_num, column=col_num).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        adjusted_width = min(max_length + 2, 30)
        summary_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Формируем имя файла
    filename = f"klassy_s_uchenikami_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    
    return response